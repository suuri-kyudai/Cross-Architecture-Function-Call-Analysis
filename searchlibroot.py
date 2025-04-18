import sys
import openpyxl
from processargs import Pkfile

# IDAで逆アセンブルして作成したリストファイルから
# symlistにある関数名の中で呼び出されている時系列情報を削除
# つまりライブラリ関数をルートとする連結成分を削除
# 使うならlist2nx.pyの前にかますこと!

wb = openpyxl.load_workbook('path/to/libraryfunc_list.xlsx')
# 最終行
maxrow = sheet.max_row + 1


# 関数名symがユーザー定義関数かどうかを判定する関数
# エクセルシートにのっているシンボルはライブラリ関数として判定する
def is_userfunc(sym):
    for row in range(1, maxrow):
        libsym = sheet.cell(row=row, column=1).value
        if libsym == sym:
            return False
    else:
        return True


if __name__ == "__main__":
    input_path_list = sys.argv[2:]
    output_path = sys.argv[1]
    all_len_func_list = 0
    all_len_func_list_nocall_fromlib = 0
    for input_path in input_path_list:
        pkfile = Pkfile(input_path, output_path)
        # func_list = [['node_name', ['node_name1', 'node_name2', ...]]]
        func_list = pkfile.load(pkfile.path)
        # pkfileがtruncateしていたらfunc_listに何もはいらない
        if func_list:
            nocall_fromlib = [f for f in func_list if is_userfunc(f[0])]
            all_len_func_list += len(func_list)
            all_len_func_list_nocall_fromlib += len(nocall_fromlib)
            print(f"{pkfile.name}:", end=" ")
            print(len(func_list), len(nocall_fromlib))
            pkfile.suffix = "rmlibtree"
            pkfile.extension = ".pk"
            pkfile.setoutput()
            pkfile.dump(pkfile.output, nocall_fromlib)
    else:
        print(all_len_func_list / len(input_path_list),
              all_len_func_list_nocall_fromlib / len(input_path_list))
